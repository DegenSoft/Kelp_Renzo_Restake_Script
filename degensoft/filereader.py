import csv
import random

from openpyxl import load_workbook

from degensoft.decryption import is_base64, decrypt_private_key
from degensoft.utils import load_lines


def load_and_decrypt_wallets(filename, password=''):
    """
    Will load wallets.txt file, if wallets was encrypted, trying to decrypt with the password provided
    :param filename:
    :param password:
    :param shuffle:
    :return:
    """
    lines = load_lines(filename)
    wallets = []
    for line in lines:
        if password and is_base64(line):
            wallets.append(decrypt_private_key(line, password))
        else:
            wallets.append(line)
    return wallets


class FileReader:
    def __init__(self, file_name):
        self.wallets = []
        self.file_name = file_name

    def load(self) -> list:
        raise NotImplemented()

    def decrypt(self, password):
        for item in self.wallets:
            for key in item:
                if is_base64(item[key]):
                    item[key] = decrypt_private_key(item[key], password)

    def is_encrypted(self):
        for item in self.wallets:
            for key in item:
                if is_base64(item[key]):
                    return True
        return False

    def check(self) -> bool:
        return True


class CsvFileReader(FileReader):
    def load(self) -> list:
        with open(self.file_name, 'r') as f:
            return self.load_csv(f)

    def load_csv(self, stream) -> list:
        dialect = csv.Sniffer().sniff(stream.readline(), delimiters=";,")
        stream.seek(0)
        reader = csv.DictReader(stream, dialect=dialect)
        for row in reader:
            self.wallets.append(row)
        return self.wallets


class XlsxFileReader(FileReader):
    def load(self) -> list:
        with open(self.file_name, 'r') as f:
            return self.load_xlsx(f)

    def load_xlsx(self, stream) -> list:
        workbook = load_workbook(filename=stream)
        sheet = workbook.worksheets[0]
        columns = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            self.wallets.append(dict(zip(columns, row)))
        return self.wallets


class UniversalFileReader(XlsxFileReader, CsvFileReader, FileReader):
    def load(self):
        if self.file_name.endswith('.xlsx'):
            with open(self.file_name, 'rb') as f:
                return self.load_xlsx(f)
        else:
            with open(self.file_name, 'r') as f:
                return self.load_csv(f)
